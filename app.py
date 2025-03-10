from flask import Flask, request, Response, jsonify
import pandas as pd
from flask_cors import CORS, cross_origin
from datetime import datetime
import os
import json
from datetime import datetime
from io import BytesIO
import functions as fn
import requests
from PIL import Image
from openpyxl.workbook import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from flask_api import status
from openpyxl.worksheet.datavalidation import DataValidation
import gspread
import logging
import sys
import time
import threading
import uuid
import redis
from dotenv import load_dotenv
from werkzeug.serving import WSGIRequestHandler
from pathlib import Path

# Suppress Werkzeug logs
logging.getLogger("werkzeug").setLevel(logging.ERROR)

load_dotenv()

pd.options.mode.chained_assignment = None  # default='warn'

app = Flask(__name__)

app.config["SECRET_KEY"] = os.urandom(28)
# app.config["HOSTNAME"] = os.getenv("hostname")
# app.config["USERNAME"] = os.getenv("username")
# app.config["PASSWORD"] = os.getenv("password")
app.config["GSHEETSKEY"] = os.getenv("gsheetskey")
app.config["ca_auth_token"] = os.getenv("ca_auth_token")
app.config["ca_refresh_token"] = os.getenv("ca_refresh_token")
app.config["redis_password"] = os.getenv("redis_password")
dev = os.getenv("dev")

CORS(app, supports_credentials=True, resources={r"/*": {"origins": "*"}})

if dev == "False":
    print("hello")
    # Configure Redis connection
    redis_client = redis.Redis(
        host="127.0.0.1", port=6379, password=app.config["redis_password"]
    )
else:
    print("world")
    # Configure Redis connection
    redis_client = redis.Redis(host="localhost", port=6379)


def update_task_field(task_id, field, value):
    # Retrieve the current task data
    task_data = redis_client.get(task_id)
    if task_data:
        task_data = json.loads(task_data)  # Convert JSON string to dictionary
    else:
        task_data = {}  # Initialize if no existing data

    # Update the specified field
    task_data[field] = value

    # Save the updated dictionary back to Redis
    redis_client.set(task_id, json.dumps(task_data))


# logging.basicConfig(filename='DebugLogs.log', encoding='utf-8', level=logging.DEBUG)
# logger = logging.getLogger(__name__)

# Set the logging level to DEBUG so that it logs all messages.
logger = app.logger
logger.setLevel(logging.DEBUG)

# Create a log file and configure the file handler.
log_handler = logging.FileHandler("app.log")
log_handler.setLevel(logging.DEBUG)

# Create a formatter to format log messages.
formatter = logging.Formatter("%(asctime)s - %(name)s - %(levelname)s - %(message)s")
log_handler.setFormatter(formatter)

# Add the file handler to the app's logger.
logger.addHandler(log_handler)


def handle_exception(exc_type, exc_value, exc_traceback):
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return

    app.logger.error(
        "Uncaught exception", exc_info=(exc_type, exc_value, exc_traceback)
    )


sys.excepthook = handle_exception


@app.route("/", methods=["GET"])
def index():
    json = {"status": 200, "message": "This is a running API!"}
    return jsonify(json)


# used
## endpoint for uploading an elastic listing file
@app.route("/ListingUpload", methods=["POST", "GET"])
def ListingUpload():
    if request.method == "GET":
        return "Upload a file using the API"

    # post
    file = request.files["file"]
    elasticBool = request.form["elasticBool"]

    # Save the file locally
    file.save("uploaded_file.xlsx")

    # Read the Excel file into a Pandas DataFrame
    df = pd.read_excel("uploaded_file.xlsx")
    if "Primary Category" not in df.columns:
        error = "Primary Category column not found"
        return error, status.HTTP_400_BAD_REQUEST
    if elasticBool == "true":
        # SalesForce
        SalesForceDf = pd.read_excel("/Users/wilverine7/Desktop/Salesforce.xlsx")

        for x in range(len(df)):
            if df["Color Name"][x] == "No Color":
                df.at[x, "Color Name"] = ""
            if df["Size"][x] == "1SZ":
                df.at[x, "Size"] = ""
            if "Category" in df.columns:
                if df["Category"][x] == "Skis" and df["Size"][x] != "":
                    df.at[x, "Size"] = f"{df['Size'][x]}cm"
            PONumber = df["PO #"][x]
            PONumber = PONumber.split("-")[0]
            df.at[x, "PO #"] = PONumber

        # build new df that has all the columns we want for sku creation
        newDf = pd.DataFrame(
            columns=[
                "Existing Sku",
                "SKU",
                "Existing Parent SKU",
                "Title",
                "Manufacturer Title",
                "Model Name",
                "Primary Color",
                "Config Color",
                "Size",
                "Gender",
                "Sport",
                "Primary Category",
                "Classification",
                "Brand",
                "Quantity",
                "Cost",
                "Wholesale Cost",
                "Sale Price",
                "Retail Price",
                "MAP Price",
                "Price LS",
                "UPC",
                "EAN",
                "MPN",
                "Supplier",
                "Supplier Part Number",
                "PO Number",
                "Sales Opportunity",
                "MAP Restrictions",
                "Selling Channels",
                "Territories",
                "Inventory",
                "Type",
                "Reorder Point",
                "Model Year",
            ]
        )
        # all data that gets moved over simply just copy over
        newDf["PO Number"] = df["PO #"]
        newDf["Sales Opportunity"] = df["PO #"]
        newDf["Model"] = df["Style Name"]
        newDf["MPN"] = df["SKU"]
        newDf["Part Number"] = df["SKU"]
        newDf["Wholesale Cost"] = df["Wholesale Price"]
        newDf["Retail Price"] = df["Retail Price"]
        newDf["Quantity"] = df["Quantity Requested"]
        newDf["Color"] = df["Color Name"]
        newDf["Size"] = df["Size"]
        newDf.fillna("", inplace=True)

        # Check Length <=10 then delete, 11 =12 UPC, 13-15 EAN, > 15 Delete, IF (ISNUMBER=FALSE, DELETE)
        for x in range(len(df)):
            if len(str(df["UPC"][x])) > 15 or len(str(df["UPC"][x])) <= 10:
                newDf.at[x, "UPC"] = df["UPC"][x].astype(str)
                newDf.at[x, "EAN"] = df["UPC"][x].astype(str)

            elif len(str(df["UPC"][x])) >= 13:
                newDf.at[x, "EAN"] = df["UPC"][x].astype(str)
                newDf.at[x, "UPC"] = ""

            elif len(str(df["UPC"][x])) >= 11:
                newDf.at[x, "UPC"] = df["UPC"][x].astype(str)
                newDf.at[x, "EAN"] = ""

        # convert unisex to Men's, Women's
        for x in range(len(df)):
            string = df["Gender"][x]
            if string.find("Unisex") != -1:
                string = string.replace("Unisex", "Men's, Women's")
                df.at[x, "Gender"] = string
            df["Gender"][x]
            newDf["Gender"] = df["Gender"]

        PO = newDf["PO Number"][0]

        for x in range(len(SalesForceDf["Sales Opp. Number"])):
            string = SalesForceDf["Sales Opp. Number"][x]
            N = 5
            # get length of string
            length = len(string)

            # create a new string of last N characters
            Str2 = "L9" + string[length - N :]
            SalesForceDf.at[x, "PO Number"] = Str2
            if Str2 == PO:
                print("found")
                newDf["MAP Restrictions"] = SalesForceDf.at[x, "MAP Restrictions"]
                newDf["Supplier"] = SalesForceDf.at[x, "Supplier Name"]
                newDf["Brand"] = SalesForceDf.at[x, "Supplier.Brands"]

                # territroy restrictions
                if SalesForceDf["Territory Restrictions"][x] == "3 - No Restrictions":
                    newDf["Territories"] = (
                        "Argentina;Australia;Chile;Ireland;Japan;New Zealand;UK;United States;Canada"
                    )
                elif SalesForceDf["Territory Restrictions"][x] == "2 - Other":
                    newDf["Territories"] = "United States; Canada"
                elif SalesForceDf["Territory Restrictions"][x] == "1 - US Only":
                    newDf["Territories"] = "United States"
        newDf["Primary Category"] = df["Primary Category"]
        primaryDf = newDf.copy()
    else:
        primaryDf = df.copy()

    # this is the category to attribute set sheet
    # open Attribute Set Attribute sheet and convert it to a DF
    attributeUrl = "https://docs.google.com/spreadsheets/d/1tXm039Fcj16Qn1rWd6HzpyoqQ_l0H64tlmV_0nVDmIk/edit#gid=626682809"
    sa = gspread.service_account_from_dict(app.config["GSHEETSKEY"])
    sh = sa.open_by_url(attributeUrl)
    ws = sh.worksheet("PrimaryToAttributeSet")
    primaryToAttributeDf = pd.DataFrame(ws.get_all_records())

    # this is the sheet with attribute values that will populate dropdowns
    ws = sh.worksheet("Attribute Values")
    attributeValuesDf = pd.DataFrame(ws.get_all_records())

    # create a copy of the original sheet
    wb = Workbook()
    openpyxl_ws = wb.active
    copySheet = wb.create_sheet(title="CopyOfOriginal")
    for r in dataframe_to_rows(primaryDf, index=False, header=True):
        copySheet.append(r)

    # drop rows where Primary Category is null
    primaryDf = primaryDf.dropna(subset=["Primary Category"])

    # get the categories from the listing sheet
    categories = primaryDf["Primary Category"].unique()

    # maxx keeps track of the total attribute name columns added so we can know how many attribute value columns we need to add
    maxx = 0
    # build the attribute set columns in the listing sheet
    for category in categories:
        # get just the row that matches the category
        attributeDf = primaryToAttributeDf[
            primaryToAttributeDf["Primary Category"] == category
        ]
        attributeDf = attributeDf.replace("", pd.NA)
        attributeDf = attributeDf.dropna(axis=1)
        x = 1

        # loop through the columns and get the attribute names
        attributeValueList = []
        for columnName in attributeDf.columns:
            if columnName == f"attribute_id {x}":
                # get the attribute name
                name = attributeDf[columnName].iloc[0]

                # add the attribute name to the listing sheet
                primaryDf.loc[
                    primaryDf["Primary Category"] == category, [f"Attribute{x}Name"]
                ] = name
                x += 1
                if maxx < x:
                    maxx = x

    # convert primarydf to openpyxl, get the column header and if it matches f"Attribute{x}Name"
    # then get the cell below. Add a new column next to it that is datavalidation with the values coming from the attribute values sheet
    for r in dataframe_to_rows(primaryDf, index=False, header=True):
        openpyxl_ws.append(r)
    x = 1
    maxcol = openpyxl_ws.max_column
    columnNumber = 1
    NewSheet = wb.create_sheet(title="attributeValue")
    attribute = ""
    for column in range(1, maxcol + maxx + 1):
        cell = openpyxl_ws.cell(row=1, column=column)

        if cell.value == f"Attribute{x}Name":
            openpyxl_ws.insert_cols(column + 1)
            cell = openpyxl_ws.cell(row=1, column=column + 1)
            cell.value = f"Attribute{x}Value"
            row = 2
            while row <= openpyxl_ws.max_row:
                # for row in range(2, openpyxl_ws.max_row + 1):
                attribute = openpyxl_ws.cell(row=row, column=column)
                print(attribute.value)

                # if the cell before is the same as the current cell then we can reuse the datavalidation formula
                if (
                    attribute.value
                    == openpyxl_ws.cell(row=row - 1, column=column).value
                ):
                    dv = DataValidation(
                        type="list",
                        formula1=formula,
                        allow_blank=True,
                    )
                    # Optionally set a custom error message
                    dv.error = "Your entry is not in the list"
                    dv.errorTitle = "Invalid Entry"
                    # Optionally set a custom prompt message
                    dv.prompt = "Please select from the list"
                    dv.promptTitle = "List Selection"

                    # add the data validation to only the cell next to the current cell
                    cell = openpyxl_ws.cell(row=row, column=column + 1)
                    openpyxl_ws.add_data_validation(dv)
                    dv.add(cell)
                    row += 1

                # if the cells don't match we need to get new values to build the data validation
                else:
                    # print(attribute.value)
                    filteredDf = attributeValuesDf[
                        attributeValuesDf["name"] == attribute.value
                    ]
                    if not filteredDf.empty:
                        # print(filteredDf)
                        filteredDf.replace("", pd.NA, inplace=True)
                        filteredDf.dropna(axis=1, inplace=True)
                        count = 1
                        attributeValueList = []
                        for columnName in filteredDf.columns:
                            if columnName == f"attribute_value {count}":
                                attributeValueList.append(
                                    filteredDf[columnName].iloc[0]
                                )

                                count += 1
                        print(attributeValueList)

                        for printRow in range(len(attributeValueList)):
                            cell = NewSheet.cell(row=printRow + 1, column=columnNumber)
                            cell.value = attributeValueList[printRow]
                            maxRow = printRow + 1

                        ColumnLetter = openpyxl_ws.cell(
                            row=1, column=columnNumber
                        ).column_letter
                        columnNumber = columnNumber + 1
                        formula = (
                            f"attributeValue!{ColumnLetter}1:{ColumnLetter}{maxRow}"
                        )
                        dv = DataValidation(
                            type="list",
                            formula1=formula,
                            allow_blank=True,
                        )
                        # Optionally set a custom error message
                        dv.error = "Your entry is not in the list"
                        dv.errorTitle = "Invalid Entry"
                        # Optionally set a custom prompt message
                        dv.prompt = "Please select from the list"
                        dv.promptTitle = "List Selection"

                        # add the data validation to only the cell next to the current cell
                        cell = openpyxl_ws.cell(row=row, column=column + 1)
                        openpyxl_ws.add_data_validation(dv)
                        dv.add(cell)
                        row += 1

                    else:
                        # if the attribute isn't in the attribute value sheet then just add a blank data validation
                        formula = ""
                        dv = DataValidation(
                            type="list",
                            formula1=formula,
                            allow_blank=True,
                        )
                        # Optionally set a custom error message
                        dv.error = "Your entry is not in the list"
                        dv.errorTitle = "Invalid Entry"
                        # Optionally set a custom prompt message
                        dv.prompt = "Please select from the list"
                        dv.promptTitle = "List Selection"

                        # add the data validation to only the cell next to the current cell
                        cell = openpyxl_ws.cell(row=row, column=column + 1)
                        openpyxl_ws.add_data_validation(dv)
                        dv.add(cell)

                        print("no match")
                        row += 1

            x += 1
    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)  # go to the beginning of the stream

    return Response(
        excel_stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# used
# Endpoint to upload a single image from the url
@app.route("/UrlUpload", methods=["POST"])
def UrlUpload():
    app.logger.info("UrlUpload")
    if request.form["url"] == "":
        imageFile = request.files["file"]
        imagePath = ""
    else:
        sep = "?"
        imagePath = request.form["url"]
        imagePath = imagePath.split(sep, 1)[0]
        r = requests.get(imagePath, stream=True)
        if r.status_code != 200:
            imagePath = request.form["url"]

    sku = request.form["sku"]
    sku.replace(" ", "")
    imgNum = request.form["imageNumber"]
    flag = request.form["flag"] == "true"
    remBg = request.form["removeBackground"] == "true"
    imageName = f"{sku}_Img{imgNum}"

    folder_name = datetime.today().strftime("%Y-%m-%d")
    # creates a variable to pass to the html page to display the image and url
    BikeWagonUrl = f"https://l9golf.com/images/media/L9/{folder_name}/{imageName}.jpg"
    server_path = f"/var/www/images/media/L9/{folder_name}/{imageName}.jpg"
    server_dir = f"/var/www/images/media/L9/{folder_name}"

    if os.path.isfile(server_path) and flag == False:
        flag = True
        error = "Duplicate Image. Would you like to overwrite the image?"
        # displayImage = (
        #     f"https://l9golf.com/images/media/L9/{folder_name}/{imageName}.jpg"
        # )
        data = {
            "error": error,
            "flag": flag,
            "displayImage": BikeWagonUrl,
        }
        return data
    else:
        if not os.path.exists(server_dir):
            os.makedirs(server_dir)
    if imagePath == "":
        # handle the file upload
        image = Image.open(imageFile).convert("RGBA")
        if remBg:
            image_io = fn.removeBackground(image)
        else:
            image_io = fn.process_image(image)

        # Now save the file
        with open(server_path, "wb") as f:
            f.write(image_io.getvalue())

        data = {"displayImage": BikeWagonUrl, "flag": False}
        return data, 200
    else:
        # handle the url upload
        try:
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
            }
            # open the image from the url
            response = requests.get(imagePath, stream=True, headers=headers)
            # if the user wants to remove background it processes here.
            if remBg:
                image = Image.open(BytesIO(response.content))
                image_io = fn.removeBackground(image)

            else:
                image = Image.open(BytesIO(response.content)).convert("RGBA")
                # process the image by passing PIL image to the function
                image_io = fn.process_image(image)
            # Now save the file
            with open(server_path, "wb") as f:
                f.write(image_io.getvalue())

            data = {"displayImage": BikeWagonUrl, "flag": False}

            return data, 200
        except:
            error = "Invalid URL"
            # if the image wouldn't open then the url is invalid
            json = {"error": error}
            app.logger.error(f"Invalid URL: {error}")
            return json


@app.route("/CaUpload", methods=["POST"])
@cross_origin(supports_credentials=True)
def CaUpload():
    app.logger.info("CaUpload")
    ca_auth_token = app.config["ca_auth_token"]
    ca_refresh_token = app.config["ca_refresh_token"]
    errors = []
    uploadSuccess = []
    try:
        ca_auth_token = fn.getToken(ca_refresh_token, ca_auth_token)
        if ca_auth_token.startswith("Request failed"):
            raise Exception
    except:
        app.logger.error(ca_auth_token)
        return ca_auth_token, 500

    clientUrl = request.form["clientUrl"]
    if clientUrl == "urlUpload":
        imageUrl = request.form["url"]
        sku = request.form["sku"]
        imageNum = request.form["imageNumber"]
        response = fn.caUpload(sku, imageUrl, imageNum, ca_auth_token)
        if response == "success":
            uploadSuccess = sku
        else:
            errors = sku
    elif clientUrl == "csvUpload":
        downloadWithErrors = request.form["downloadWithErrors"]
        # gets the df formatted in the input format and converts it to Channel Advisor format
        df = request.form["df"]
        df = pd.read_json(df, orient="index")
        print(df)

        if downloadWithErrors == "true":
            df = df.fillna("")
        else:
            try:
                errorDict = request.form["errorDict"]
                errorDict = json.loads(errorDict)
                if errorDict != {}:
                    for key in errorDict:
                        df = df[df["PARENT_SKU_COLOR"] != key]
            except:
                error = "Select the download with errors box and try again."
                return Response(error, status.HTTP_400_BAD_REQUEST)
        uploadCount = 0
        for sku in df.index:
            x = 1
            dfSku = df[df.index == sku]
            dfSku = dfSku.dropna(axis=1, how="all")
            while f"Server Image {x}" in dfSku.columns:
                imageUrl = dfSku.loc[sku, f"Server Image {x}"]
                if uploadCount >= 750:
                    # wait 30 seconds
                    time.sleep(30)
                    print("waiting")
                response = fn.caUpload(sku, imageUrl, x, ca_auth_token)
                if response != "success":
                    # handle errors
                    errors.append({"SKU": sku, "imgNum": x})
                    app.logger.error(f"unable to upload: {sku}: image number {x}")
                else:
                    uploadSuccess.append(sku)
                    app.logger.info(f"succesfully uploaded: {sku}")
                uploadCount += 1
                x += 1
    responseJson = {"errors": errors, "success": uploadSuccess}
    return responseJson, 200


def ImageCsv(task_id, file, folder):
    app.logger.info("ImageCsv - POST")

    df = pd.read_csv(file)

    # if the url doesn't work, keep track of it and remove it from the df
    BrokenUrlDict = {}

    df.columns = map(str.upper, df.columns)
    df.columns = df.columns.str.strip()
    df.columns = df.columns.str.replace(" ", "_")

    df.dropna(subset=["IMAGE_1"], inplace=True)
    folder_name = datetime.today().strftime("%Y-%m-%d")

    chunk_size = 50
    num_chunks = len(df) // chunk_size + (len(df) % chunk_size > 0)

    if ("PARENT_SKU_COLOR" in df.columns) and (df["PARENT_SKU_COLOR"].notnull().any()):
        columnIdentifier = "PARENT_SKU_COLOR"
    else:
        columnIdentifier = "SKU"

    # allows you to upload a file or url
    # doesn't require the export sheet. You can export the sourcing sheet
    server_dir = f"/var/www/images/media/L9/{folder_name}"

    for i in range(num_chunks):
        app.logger.info(f"processing chunk {i+1}")
        chunk = df.iloc[i * chunk_size : (i + 1) * chunk_size]

        df_copy = chunk.dropna(axis=1, how="all")

        # maxPictureCount is used to extend the df columns to the right number of images.
        maxImageColCount = 1

        # see how many images columns there are and add one extra to avoid index out of range error
        while f"IMAGE_{maxImageColCount}" in df_copy.columns:
            maxImageColCount += 1
        maxImageColCount -= 1

        columns = []
        if columnIdentifier == "PARENT_SKU_COLOR":
            chunk["PARENT_SKU_COLOR"] = chunk["PARENT_SKU_COLOR"].astype(str)
            uniqueCombo = chunk["PARENT_SKU_COLOR"].unique()
            print(uniqueCombo[0])
            if len(uniqueCombo) == 1 and (
                uniqueCombo[0] == "" or uniqueCombo[0] == "nan"
            ):
                columnIdentifier = "SKU"
                uniqueCombo = chunk["SKU"].unique()

            totalUploaded = 0
            totalImages = (
                chunk[
                    [
                        "IMAGE_1",
                        "IMAGE_2",
                        "IMAGE_3",
                        "IMAGE_4",
                        "IMAGE_5",
                        "IMAGE_6",
                        "IMAGE_7",
                        "IMAGE_8",
                        "IMAGE_9",
                    ]
                ]
                .nunique()
                .sum()
            )
            print(f"total images: {totalImages}")
            update_task_field(
                task_id=task_id,
                field="progress",
                value=(totalUploaded / totalImages),
            )
            update_task_field(
                task_id=task_id, field="chunks", value=f"{i+1}/{num_chunks}"
            )
        else:
            uniqueCombo = chunk["SKU"].unique()
            totalUploaded = 0
            totalImages = (
                chunk[
                    [
                        "IMAGE_1",
                        "IMAGE_2",
                        "IMAGE_3",
                        "IMAGE_4",
                        "IMAGE_5",
                        "IMAGE_6",
                        "IMAGE_7",
                        "IMAGE_8",
                        "IMAGE_9",
                    ]
                ]
                .count()
                .sum()
            )
            print(f"total images: {totalImages}")
            update_task_field(
                task_id=task_id,
                field="progress",
                value=(totalUploaded / totalImages),
            )
            update_task_field(
                task_id=task_id, field="chunks", value=f"{i+1}/{num_chunks}"
            )

        if not os.path.exists(server_dir):
            os.makedirs(server_dir)
            app.logger.info("Created new folder")
        try:
            # getting the uniqueSku problem is you download images multiple times
            for combo in uniqueCombo:
                app.logger.debug(f"Processing combo: {combo}")
                urlList = ""

                # x keeps track of the number of images for each parent SKU color combo
                x = 1
                # CaDf.append([{"Inventory Number": sku}])
                dfCombo = chunk[chunk[columnIdentifier] == combo]

                # if a parent_color combo has more than one unique URL in the comboDf we need to handle it differently
                uniquePath = dfCombo[f"IMAGE_{x}"].unique()
                # dfCombo.dropna(axis=1, inplace=True)
                dfCombo.reset_index(drop=True, inplace=True)
                # print(dfCombo)
                # error catch: Could also change this to process the filtered df by Child sku and not make the user do it manually
                # Allows there to be unique urls even if the parent sku combo is the same

                ### ASSUMPTION ####
                # if the Image_1 url is the same for all rows of the parent sku combo then it will process them all together.
                # if we there is a case where image 1 is the same but image 2 is different that is not handled.
                if len(uniquePath) > 1:
                    # this goes by SKU rather than combo so if there are multiple unique urls for a sku it will process them
                    print(uniquePath)
                    for unique in uniquePath:
                        # reset to the original dfCombo
                        dfCombo = chunk[chunk[columnIdentifier] == combo]
                        x = 1
                        # get each line with unique URLS
                        dfCombo = dfCombo[dfCombo[f"IMAGE_{x}"] == unique]
                        dfCombo.reset_index(drop=True, inplace=True)
                        sku = dfCombo["SKU"][0]
                        print(dfCombo["SKU"][0])

                        print(dfCombo[f"IMAGE_{x}"][0])
                        while (
                            f"IMAGE_{x}" in dfCombo.columns
                            and dfCombo[f"IMAGE_{x}"].count() > 0
                        ):
                            if (
                                x == 1
                                and str(dfCombo["SKI_FLIP"][0] or "").upper() == "TRUE"
                            ):
                                flip = True
                            else:
                                flip = False
                            # if it is a url
                            imageUrl = dfCombo[f"IMAGE_{x}"][0]
                            sep = "?"
                            imageUrl = imageUrl.split(sep, 1)[0]

                            try:
                                print(imageUrl)
                                r = requests.get(imageUrl, stream=True)
                                if r.status_code != 200:
                                    imageUrl = dfCombo[f"IMAGE_{x}"][0]
                                else:
                                    requests.get(imageUrl, stream=True)

                                server_path = f"/var/www/images/media/L9/{folder_name}/{sku}_{x}.jpg"

                                try:
                                    headers = {
                                        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
                                    }
                                    response = requests.get(
                                        imageUrl, stream=True, headers=headers
                                    )
                                    image = Image.open(
                                        BytesIO(response.content)
                                    ).convert("RGBA")
                                    if flip:
                                        builtSki = fn.skiBuilder(imageUrl)
                                        # Save the resized image to a file-like object
                                        image_io = BytesIO()
                                        builtSki.convert("RGB").save(image_io, "JPEG")

                                        image_io.seek(0)
                                    else:
                                        image_io = fn.process_image(image)
                                    with open(server_path, "wb") as f:
                                        f.write(image_io.getvalue())
                                    totalUploaded += 1
                                    app.logger.info(
                                        f"Total images uploaded: {totalUploaded}"
                                    )
                                    progress = totalUploaded / totalImages
                                    if progress == 1:
                                        progress = 0.99
                                    update_task_field(
                                        task_id=task_id,
                                        field="progress",
                                        value=progress,
                                    )
                                    BikeWagonUrl = f"https://l9golf.com/images/media/L9/{folder_name}/{sku}_{x}.jpg"

                                    # adds a column with the value of the new url to the df
                                    chunk.loc[
                                        chunk["SKU"] == sku,
                                        f"Server Image {x}",
                                    ] = BikeWagonUrl

                                    # adds the new column name to the list so we can get all the needed columns later
                                    if f"Server Image {x}" not in columns:
                                        columns.append(f"Server Image {x}")

                                    # adds the url to the urlList variable so we can add it to the csv file for Channal Advisor upload
                                    if urlList == "":
                                        urlList = BikeWagonUrl
                                    else:
                                        urlList = urlList + "," + BikeWagonUrl

                                except Exception as e:
                                    app.logger.error(f"Error: {str(e)}")
                                    print(f"Error: {str(e)}")
                                    if sku not in BrokenUrlDict:
                                        BrokenUrlDict[sku] = f"IMAGE_{x}"
                                    else:
                                        BrokenUrlDict[sku] += f", IMAGE_{x}"

                                x += 1

                            except:
                                # this will be the image name in the folder that is uploaded
                                imagePath = dfCombo[f"IMAGE_{x}"][0]

                                # if the imagePath contains a . split the string and get everything before the .
                                if "." in imagePath:
                                    fileName = imagePath.split(".")[0]
                                    imagePath = fileName
                                else:
                                    fileName = imagePath

                                for filename, byte_data in folder.items():
                                    if (
                                        filename.endswith(".jpg")
                                        or filename.endswith(".png")
                                        or filename.endswith(".jpeg")
                                        or filename.endswith(".webp")
                                        or filename.endswith(".JPG")
                                        or filename.endswith(".JPEG")
                                        or filename.endswith(".PNG")
                                        or filename.endswith(".WEBP")
                                    ):
                                        imageName = filename.rsplit("/", 1)[-1]
                                        # remove the file extenstion from the imageName
                                        imageName = imageName.split(".")[0]

                                        if imageName == fileName:
                                            imagePath = byte_data
                                server_path = f"/var/www/images/media/L9/{folder_name}/{sku}_{x}.jpg"
                                try:
                                    image = Image.open(imagePath).convert("RGBA")
                                    if flip:
                                        builtSki = fn.skiBuilder(imagePath)
                                        # Save the resized image to a file-like object
                                        image_io = BytesIO()
                                        builtSki.convert("RGB").save(image_io, "JPEG")

                                        image_io.seek(0)
                                    else:
                                        image_io = fn.process_image(image)

                                    with open(server_path, "wb") as f:
                                        f.write(image_io.getvalue())
                                    totalUploaded += 1
                                    app.logger.info(
                                        f"Total images uploaded: {totalUploaded}"
                                    )
                                    progress = totalUploaded / totalImages
                                    if progress == 1:
                                        progress = 0.99
                                    update_task_field(
                                        task_id=task_id,
                                        field="progress",
                                        value=progress,
                                    )
                                    BikeWagonUrl = f"https://l9golf.com/images/media/L9/{folder_name}/{sku}_{x}.jpg"
                                    chunk.loc[
                                        chunk[columnIdentifier] == sku,
                                        f"Server Image {x}",
                                    ] = BikeWagonUrl
                                    if f"Server Image {x}" not in columns:
                                        columns.append(f"Server Image {x}")
                                    if urlList == "":
                                        urlList = BikeWagonUrl
                                    else:
                                        urlList = urlList + "," + BikeWagonUrl

                                except Exception as e:
                                    app.logger.warn(f"Error: {str(e)} -- {imagePath}")
                                    print(imagePath)
                                    print(f"Error: {str(e)}")
                                    if sku not in BrokenUrlDict:
                                        BrokenUrlDict[sku] = f"IMAGE_{x}"
                                    else:
                                        BrokenUrlDict[sku] += f", IMAGE_{x}"
                                    print(BrokenUrlDict)

                                x += 1

                else:
                    # this process the df using just parent so if all children have the same url it will process them
                    while (
                        f"IMAGE_{x}" in dfCombo.columns
                        and dfCombo[f"IMAGE_{x}"].count() > 0
                    ):
                        if (
                            x == 1
                            and str(dfCombo["SKI_FLIP"][0] or "").upper() == "TRUE"
                        ):
                            flip = True
                        else:
                            flip = False
                        ####### I need to fix x and make sure the variable isn't reused####

                        # if the first row doesn't have an image but another row does have an image we need to use that image

                        # if it is a url
                        sep = "?"
                        imageUrl = dfCombo[f"IMAGE_{x}"][0]
                        imageUrl = imageUrl.split(sep, 1)[0]
                        try:
                            r = requests.get(imageUrl, stream=True)
                        except:
                            status_code = 500
                        else:
                            status_code = r.status_code
                        if status_code != 200:
                            imageUrl = dfCombo[f"IMAGE_{x}"][0]
                        if imageUrl == "" or pd.isnull(imageUrl):
                            dfCombo = dfCombo[dfCombo[f"IMAGE_{x}"] != ""]
                            dfCombo = dfCombo[dfCombo[f"IMAGE_{x}"].notnull()]
                            dfCombo.reset_index(drop=True, inplace=True)
                            imageUrl = dfCombo[f"IMAGE_{x}"][0]
                        try:
                            headers = {
                                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
                                # "Referer": "https://your-referer-site.com",  # Replace with the actual referer if needed
                            }
                            response = requests.get(
                                imageUrl, stream=True, headers=headers
                            )
                        except:
                            # this is a broken url so we don't get a response on purpose
                            response = requests.get(
                                "https://l9golf.com/images/BrokenUrl",
                                stream=True,
                            )

                        server_path = (
                            f"/var/www/images/media/L9/{folder_name}/{combo}_{x}.jpg"
                        )
                        app.logger.info(f"{imageUrl} -- {response.status_code}")
                        if response.status_code == 200:
                            try:
                                image = Image.open(BytesIO(response.content)).convert(
                                    "RGBA"
                                )
                                if flip:
                                    builtSki = fn.skiBuilder(imageUrl)
                                    # Save the resized image to a file-like object
                                    image_io = BytesIO()
                                    builtSki.convert("RGB").save(image_io, "JPEG")

                                    image_io.seek(0)
                                else:
                                    image_io = fn.process_image(image)
                                with open(server_path, "wb") as f:
                                    f.write(image_io.getvalue())
                                totalUploaded += 1
                                app.logger.info(
                                    f"Total images uploaded: {totalUploaded}"
                                )
                                progress = totalUploaded / totalImages
                                if progress == 1:
                                    progress = 0.99
                                update_task_field(
                                    task_id=task_id,
                                    field="progress",
                                    value=progress,
                                )
                                BikeWagonUrl = f"https://l9golf.com/images/media/L9/{folder_name}/{combo}_{x}.jpg"
                                chunk.loc[
                                    chunk[columnIdentifier] == combo,
                                    f"Server Image {x}",
                                ] = BikeWagonUrl
                                if f"Server Image {x}" not in columns:
                                    columns.append(f"Server Image {x}")
                                if urlList == "":
                                    urlList = BikeWagonUrl
                                else:
                                    urlList = urlList + "," + BikeWagonUrl

                            except Exception as e:
                                app.logger.error(f"Error: {str(e)}")
                                print(f"Error: {str(e)}")
                                print(imageUrl)
                                if combo not in BrokenUrlDict:
                                    BrokenUrlDict[combo] = f"IMAGE_{x}"
                                else:
                                    BrokenUrlDict[combo] += f", IMAGE_{x}"
                                print(BrokenUrlDict)

                            x += 1

                        else:
                            # this will be the image name in the folder that is uploaded
                            imagePath = dfCombo[f"IMAGE_{x}"][0]

                            # if the imagePath contains a . split the string and get everything before the .
                            if "." in imagePath:
                                fileName = imagePath.split(".")[0]
                                fileName = fileName.strip()
                            else:
                                fileName = imagePath
                                fileName = fileName.strip()

                            for filename, byte_data in folder.items():
                                imageName = filename.rsplit("/", 1)[-1]
                                # remove the file extenstion from the imageName
                                imageName = imageName.split(".")[0]

                                if imageName == fileName:
                                    imagePath = byte_data
                            server_path = f"/var/www/images/media/L9/{folder_name}/{combo}_{x}.jpg"
                            try:
                                image = Image.open(imagePath).convert("RGBA")

                                if flip == "TRUE":
                                    builtSki = fn.skiBuilder(imagePath)
                                    # Save the resized image to a file-like object
                                    image_io = BytesIO()
                                    builtSki.convert("RGB").save(image_io, "JPEG")

                                    image_io.seek(0)

                                else:
                                    image_io = fn.process_image(image)

                                with open(server_path, "wb") as f:
                                    f.write(image_io.getvalue())
                                totalUploaded += 1
                                app.logger.info(
                                    f"Total images uploaded: {totalUploaded}"
                                )
                                progress = totalUploaded / totalImages
                                if progress == 1:
                                    progress = 0.99
                                update_task_field(
                                    task_id=task_id,
                                    field="progress",
                                    value=progress,
                                )

                                BikeWagonUrl = f"https://l9golf.com/images/media/L9/{folder_name}/{combo}_{x}.jpg"
                                chunk.loc[
                                    chunk[columnIdentifier] == combo,
                                    f"Server Image {x}",
                                ] = BikeWagonUrl
                                if f"Server Image {x}" not in columns:
                                    columns.append(f"Server Image {x}")
                                if urlList == "":
                                    urlList = BikeWagonUrl
                                else:
                                    urlList = urlList + "," + BikeWagonUrl

                            except Exception as e:
                                app.logger.error(f"Error: {str(e)}")
                                app.logger.error(f"There was an issue with {imagePath}")
                                print(imagePath)
                                print(f"Error: {str(e)}")
                                if combo not in BrokenUrlDict:
                                    BrokenUrlDict[combo] = f"IMAGE_{x}"
                                else:
                                    BrokenUrlDict[combo] += f", IMAGE_{x}"
                                print(BrokenUrlDict)
                            x += 1

                chunk.loc[chunk[columnIdentifier] == combo, "Picture URLs"] = urlList

        except Exception as e:
            app.logger.error(f"ERROR: {e}")
            print(f"Error: {str(e)}")
            error = f"An error occured uploading {combo}. Please check this PARENT_SKU_COLOR and try again."
            return (error, status.HTTP_400_BAD_REQUEST)

        app.logger.info(f"Finished uploading images for chunk {i+1}")

        if "PARENT_SKU_COLOR" in chunk.columns:
            columns.extend(
                [
                    "SKU",
                    "PARENT_SKU",
                    "PARENT_SKU_COLOR",
                    "Picture URLs",
                ]
            )
        elif "PARENT_SKU" in chunk.columns:
            columns.extend(
                [
                    "SKU",
                    "PARENT_SKU",
                    "Picture URLs",
                ]
            )
        else:
            columns.extend(
                [
                    "SKU",
                    "Picture URLs",
                ]
            )
        # if there is a video column and it is not empty add video to the df we will return
        if "VIDEO" in chunk.columns and chunk["VIDEO"].count() > 0:
            chunk["Attribute1Name"] = "VideoProduct"
            chunk.rename(columns={"VIDEO": "Attribute1Value"}, inplace=True)
            columns.extend(["Attribute1Value", "Attribute1Name"])
        if "TITLE" in chunk.columns and chunk["TITLE"].count() > 0:
            columns.extend(["TITLE"])
        ServerImageColumns = []
        x = 0
        while x < maxImageColCount:
            x += 1
            columns.extend([f"IMAGE_{x}"])
            ServerImageColumns.append(f"Server Image {x}")
        try:
            chunk = chunk[columns]
        except Exception as e:
            app.logger.error(f"ERROR: {e}")
            error = "The uploaded CSV does not contain the correct columns. Please check for Title and SKU at the minimum."

            return error, status.HTTP_400_BAD_REQUEST

        # drop rows where df doesn't have an image 1 (this will get rid of skus that don't have images)
        chunk = chunk.dropna(subset=ServerImageColumns, how="all")

        try:
            chunk.set_index("SKU", inplace=True)
            csv_bytes = chunk.to_csv().encode(
                "utf-8"
            )  # Encode CSV to bytes using UTF-8
            csv_buffer = BytesIO(csv_bytes)  # Wrap the bytes into a BytesIO object

            # Reset the buffer pointer to the start
            csv_buffer.seek(0)

            dfJson = chunk.to_json(orient="index")
        except Exception as e:
            app.logger.error(f"ERROR: {e}")
            print(f"Error: {str(e)}")
            error = "The CSV either has a SKU repeated or has extra blank data. Please delete all blank rows and try again."
            return

        # create a dictionary using the sku as the key and the Server Image 1 with the url as the value

        # print(response.headers)

        # this will pass the rows as objects
        # return chunk.to_json(orient="records")

        uploadedFilesPath = "/var/www/images/media/L9/uploadedFiles"
        if not os.path.exists(uploadedFilesPath):
            os.mkdir(uploadedFilesPath)
        with open(f"{uploadedFilesPath}/{task_id}_{i+1}.csv", "wb") as f:
            f.write(csv_buffer.getvalue())

        # return jsonify(ResponseData)

    if BrokenUrlDict == {}:
        ResponseData = {"df": dfJson}

    else:
        ResponseData = {"df": dfJson, "errorDict": BrokenUrlDict}
        broken_url_json = json.dumps(BrokenUrlDict).encode("utf-8")
        json_buffer = BytesIO(broken_url_json)

        with open(
            f"/var/www/images/media/L9/uploadedFiles/{task_id}_broken_urls.json", "wb"
        ) as f:
            f.write(json_buffer.getvalue())
    update_task_field(task_id=task_id, field="progress", value=1)
    return


@app.route("/getImageCsv", methods=["POST"])
@cross_origin(supports_credentials=True)
def getImageCsv():
    task_id = request.form["task_id"]
    print(task_id)
    res = {}

    try:
        base_csv_path = f"/var/www/images/media/L9/uploadedFiles/"
        combined_df = pd.DataFrame()
        file_index = 1
        while True:
            csv_path = f"{base_csv_path}{task_id}_{file_index}.csv"
            try:
                with open(csv_path, "rb") as file:
                    df = pd.read_csv(file)
                    print(df.columns)
                    combined_df = pd.concat([combined_df, df])
                    app.logger.info(
                        f"CSV file for task {task_id} processed successfully"
                    )
                file_index += 1
            except FileNotFoundError:
                # Break the loop if no more files are found
                app.logger.info(f"No more files found after {file_index - 1} files.")
                break
        # combined_df.reset_index(drop=True, inplace=True)
        if "SKU" in combined_df.columns:
            combined_df.set_index("SKU", inplace=True)
        else:
            print(combined_df)
            combined_df.set_index("INVENTORY_NUMBER", inplace=True)

        dfJson = combined_df.to_json(orient="index")
        res["df"] = dfJson
        json_path = f"/var/www/images/media/L9/uploadedFiles/{task_id}_broken_urls.json"
        if os.path.isfile(json_path):
            try:
                with open(json_path, "r", encoding="utf-8") as json_file:
                    # Read the JSON content
                    json_data = json_file.read()  # Read the content of the file
                    broken_urls_json = json.loads(json_data)  # Parse JSON string
                    res["errorDict"] = broken_urls_json
                    app.logger.info(f"Processed broken URLs for task {task_id}")
            except Exception as e:
                app.logger.error(
                    f"Error reading or parsing JSON for task {task_id}: {str(e)}"
                )
        else:
            app.logger.error(
                f"Broken URLs file not found for task {task_id} at {json_path}"
            )

        return jsonify(res)

    except Exception as e:
        app.logger.error(f"Error processing task {task_id}: {str(e)}")
        return jsonify({"error": str(e)}), 500


@app.route("/ImageCsv", methods=["GET", "POST"])
@cross_origin(supports_credentials=True)
def start_task():
    task_id = str(uuid.uuid4())
    task_data = {"progress": 0, "chunks": "0"}
    redis_client.set(task_id, json.dumps(task_data))
    file = request.files["file"]
    folder = request.files.getlist("file[]")
    folder_data = {file.filename: BytesIO(file.read()) for file in folder}

    # Example: Print filenames and their byte size
    for filename, byte_data in folder_data.items():
        print(f"Filename: {filename}, Size: {len(byte_data.getvalue())} bytes")

    file_data = BytesIO(file.read())  # Read file content into memory

    threading.Thread(target=ImageCsv, args=(task_id, file_data, folder_data)).start()
    return jsonify({"task_id": task_id})


@app.route("/progress/<task_id>", methods=["GET"])
def get_progress(task_id):
    progress_data = redis_client.get(task_id)
    data = json.loads(progress_data)
    # app.logger.info(f"progress: ${data}")
    return jsonify(data)


# It is important that the df has the Picture URLs column or else you will just end up with a list of parent skus and the first picture from the child.
@app.route("/downloadTest", methods=["POST"])
def downloadTest():
    downloadWithErrors = request.form["downloadWithErrors"]
    # gets the df formatted in the input format and converts it to Channel Advisor format
    df = request.form["df"]
    df = pd.read_json(df, orient="index")
    folder = request.form["bool"]

    print(df)
    if downloadWithErrors == "true":
        df = df.fillna("")
    else:
        try:
            errorDict = request.form["errorDict"]
            errorDict = json.loads(errorDict)
            if errorDict != {}:
                for key in errorDict:
                    df = df[df["PARENT_SKU_COLOR"] != key]
        except:
            error = "Select the download with errors box and try again."
            return Response(error, status.HTTP_400_BAD_REQUEST)

    ###### assign parent the first image ########
    ###### add to download part of app ##########
    # if childOnly == "false":
    #     uniqueParent = df["Parent SKU"].unique()
    #     for parent in uniqueParent:
    #         UrlList = ""
    #         parentDf = df.loc[df["Parent SKU"] == parent]
    #         uniqueParentColor = parentDf["Parent SKU_Color"].unique()
    #         for combo in uniqueParentColor:
    #             ComboDf = df.loc[df["Parent SKU_Color"] == combo]
    #             url = ComboDf["Server Image 1"].iloc[0]
    #             if UrlList == "":
    #                 UrlList = url
    #             else:
    #                 UrlList = UrlList + "," + url
    #         df = pd.concat(
    #             [df, pd.DataFrame({"Picture URLs": UrlList}, index=[parent])],
    #         )
    #     print(df)

    # assigns the first image to the parent SKU
    # uniqueParent = df["PARENT_SKU"].unique()
    # for parent in uniqueParent:
    #     UrlList = ""
    #     parentDf = df.loc[df["PARENT_SKU"] == parent]
    #     uniqueParentColor = parentDf["PARENT_SKU_COLOR"].unique()
    #     for combo in uniqueParentColor:
    #         ComboDf = df.loc[df["PARENT_SKU_COLOR"] == combo]
    #         url = ComboDf["Server Image 1"].iloc[0]
    #         if UrlList == "":
    #             UrlList = url
    #         else:
    #             UrlList = UrlList + "," + url
    #     df = pd.concat(
    #         [df, pd.DataFrame({"Picture URLs": UrlList}, index=[parent])],
    #     )
    # print(df)
    df.dropna(subset=["Picture URLs"], inplace=True)
    columns = ["Picture URLs"]
    if "Attribute1Value" in df.columns:
        columns.extend(["Attribute1Name", "Attribute1Value"])

    df["Labels"] = "BigCommerce"
    ChannelAdvisorDf = df[columns]
    ChannelAdvisorDf.rename_axis("Inventory Number", inplace=True)
    csv = ChannelAdvisorDf.to_csv(index=True)
    date = datetime.now().strftime("%Y-%m-%d")

    return Response(
        csv,
        mimetype="text/csv",
        headers={"Content-disposition": f"attachment; filename={date}_ImportReady.csv"},
    )


# used for CSV page
@app.route("/DeleteImage", methods=["POST"])
def DeleteImage():
    url = request.form["url"]
    df = request.form["df"]
    df = pd.read_json(df, orient="index")
    print(df.columns)
    # df = df.replace("", pd.NA)

    for x in range(1, 10):
        if url in df[f"Server Image {x}"].values:
            # get the parent color of the row that matches the url so we can update all items with this parent color
            parentColor = df.loc[df[f"Server Image {x}"] == url, "PARENT_SKU_COLOR"][0]
            # clear the urlList from the df
            df.loc[df["PARENT_SKU_COLOR"] == parentColor, "Picture URLs"] = ""

            while f"Server Image {x+1}" in df.columns:
                # pull the row that has the index which is the sku variable above
                df.loc[df["PARENT_SKU_COLOR"] == parentColor, f"Server Image {x}"] = (
                    df.loc[
                        df["PARENT_SKU_COLOR"] == parentColor, f"Server Image {x+1}"
                    ][0]
                )
                x += 1

            else:
                df.loc[df["PARENT_SKU_COLOR"] == parentColor, f"Server Image {x}"] = ""

            break
    urlList = ""
    x = 1
    print(df.loc[df["PARENT_SKU_COLOR"] == parentColor, f"Server Image {x}"][0])
    while (
        df.loc[df["PARENT_SKU_COLOR"] == parentColor, f"Server Image {x}"][0] != ""
        and df.loc[df["PARENT_SKU_COLOR"] == parentColor, f"Server Image {x}"][0]
        != None
    ):
        print(df.loc[df["PARENT_SKU_COLOR"] == parentColor, f"Server Image {x}"][0])
        if urlList == "":
            urlList = df.loc[
                df["PARENT_SKU_COLOR"] == parentColor, f"Server Image {x}"
            ][0]
            x += 1
        else:
            urlList = (
                urlList
                + ","
                + df.loc[df["PARENT_SKU_COLOR"] == parentColor, f"Server Image {x}"][0]
            )
            x += 1
    print(urlList)
    df.loc[df["PARENT_SKU_COLOR"] == parentColor, "Picture URLs"] = urlList

    index_l9 = url.find("/L9")
    file = url[index_l9:]
    print(file)

    server_path = f"/var/www/images/media{file}"

    try:
        os.remove(server_path)
    except Exception as e:
        app.logger.error(f"ERROR: {e}")
        print(f"Error: {str(e)}")
    dfJson = df.to_json(orient="index")
    return jsonify(dfJson)


# used for single image page
@app.route("/DeleteSingleImage", methods=["POST"])
def DeleteSingleImage():
    sku = request.form["sku"]
    imageNumber = request.form["imageNumber"]
    folder_name = datetime.today().strftime("%Y-%m-%d")

    server_path = f"/var/www/images/media/L9/{folder_name}/{sku}_Img{imageNumber}.jpg"

    try:
        os.remove(server_path)
    except Exception as e:
        app.logger.error(f"Error: {e}")
        print(f"Error: {str(e)}")

    return "success"


# new packageBuilder route
@app.route("/packageBuilder", methods=["POST"])
def packageBuilder():
    imageCount = int(request.form["count"])
    packageType = request.form["type"]
    sku = request.form["sku"]
    flag = request.form["flag"]
    saveAsNew = request.form["saveAsNew"]

    if request.form["mainUrl"] == "":
        skiBoard = request.files["mainFile"]
    else:
        skiBoard = request.form["mainUrl"]

    if imageCount == 1:
        packageImage = fn.skiBuilder(skiBoard)
    elif imageCount == 2:
        if request.form["bootBindingUrl"] == "":
            bootBinding = request.files["bootBindingFile"]
        else:
            bootBinding = request.form["bootBindingUrl"]

        if packageType == "Ski":
            packageImage = fn.twoItemSkiPackageBuilder(skiBoard, bootBinding)
        elif packageType == "Board":
            packageImage = fn.twoItemBoardPackageBuilder(skiBoard, bootBinding)
    elif imageCount == 3:
        if request.form["bootUrl"] == "":
            boot = request.files["bootFile"]
        else:
            boot = request.form["bootUrl"]

        if request.form["bindingUrl"] == "":
            binding = request.files["bindingFile"]
        else:
            binding = request.form["bindingUrl"]
        if packageType == "Ski":
            packageImage = fn.skiPackageBuilder(skiBoard, boot, binding)
        elif packageType == "Board":
            packageImage = fn.boardPackageBuilder(skiBoard, boot, binding)

    image_io = BytesIO()
    packageImage.convert("RGB").save(image_io, "JPEG")

    # Upload the image to the server
    image_io.seek(0)  # Reset the file pointer to the beginning

    imageNumber = 1
    folder_name = datetime.today().strftime("%Y-%m-%d")
    server_path = f"/var/www/images/media/L9/{folder_name}/{sku}_Img{imageNumber}.jpg"
    BikeWagonUrl = (
        f"https://l9golf.com/images/media/L9/{folder_name}/{sku}_Img{imageNumber}.jpg"
    )
    server_dir = f"/var/www/images/media/L9/{folder_name}"

    try:
        # if the path exists and flag is flase then we need to tell the user that this is a duplicate.
        # if flag is true then the user already knows it is a duplicate and wants to override it
        if os.path.isfile(server_path) and flag == False:
            flag = True
            error = "Duplicate Image. Would you like to overwrite the image?"
            displayImage = f"https://l9golf.com/images/media/L9/{folder_name}/{sku}_Img{imageNumber}.jpg"
            data = {
                "error": error,
                "flag": flag,
                "displayImage": displayImage,
                "imageNumber": imageNumber,
            }
            return data
        # if the path exists and saveAsNew is true the user wants to add a new image and not override the old one
        # so we need to find the next available image number for that sku
        if os.path.isfile(server_path) and saveAsNew == "true":
            while os.path.isfile(server_path):
                imageNumber += 1
                server_path = (
                    f"/var/www/images/media/L9/{folder_name}/{sku}_Img{imageNumber}.jpg"
                )
                BikeWagonUrl = f"https://l9golf.com/images/media/L9/{folder_name}/{sku}_Img{imageNumber}.jpg"

        if not os.path.exists(server_dir):
            os.makedirs(server_dir)

        with open(server_path, "wb") as f:
            f.write(image_io.getvalue())

        data = {
            "displayImage": BikeWagonUrl,
            "flag": False,
            "error": False,
            "imageNumber": imageNumber,
        }
        return data, 200
    except Exception as e:
        print(e)

    return "success"


@app.route("/filePackageBuilder", methods=["GET", "POST"])
@cross_origin(supports_credentials=True)
def start_package_task():
    task_id = str(uuid.uuid4())
    task_data = {"progress": 0, "chunks": "0"}
    redis_client.set(task_id, json.dumps(task_data))
    file = request.files["file"]
    folder = request.files.getlist("file[]")
    file_data = BytesIO(file.read())  # Read file content into memory
    folder_data = [BytesIO(f.read()) for f in folder]  # Read each folder file content

    threading.Thread(
        target=filePackageBuilder, args=(task_id, file_data, folder_data)
    ).start()
    return jsonify({"task_id": task_id})


# @app.route("/filePackageBuilder", methods=["POST"])
def filePackageBuilder(task_id, file, folder):
    app.logger.info("filePackageBuilder - POST")
    df = pd.read_csv(file)
    # if the url doesn't work, keep track of it and remove it from the df
    BrokenUrlDict = {}
    for column in df.columns:
        print(column)

    df.columns = map(str.upper, df.columns)
    df.columns = df.columns.str.strip()
    df.columns = df.columns.str.replace(" ", "_")
    df.dropna(subset=["MAIN_IMAGE_URL"], inplace=True)
    if df["BOOT_IMAGE_URL"].count() == 0 and df["BINDING_IMAGE_URL"].count() == 0:
        test = singleSkiFileBuilder(task_id, df, app, folder)
        return test

    df = df[df["VARIATION_PARENT_SKU"] != "Parent"]
    uniqueCombo = df["VARIATION_PARENT_SKU"].unique()

    for column in df.columns:
        print(column)

    folder_name = datetime.today().strftime("%Y-%m-%d")

    columns = []
    totalUploaded = 0
    totalImages = df.shape[0]
    server_dir = f"/var/www/images/media/L9/{folder_name}"

    try:
        if not os.path.exists(server_dir):
            os.makedirs(server_dir)
        try:
            # getting the uniqueSku problem is you download images multiple times
            for combo in uniqueCombo:
                comboDf = df[df["VARIATION_PARENT_SKU"] == combo]
                sku = combo
                comboDf.reset_index(drop=True, inplace=True)
                packageType = comboDf["SKI/BOARD"][0].upper()
                if comboDf["SKI/BOARD"][0].upper() == "SKI":
                    packageType = "Ski"
                elif (
                    comboDf["SKI/BOARD"][0].upper() == "BOARD"
                    or comboDf["SKI/BOARD"][0].upper() == "SNOWBOARD"
                ):
                    packageType = "Board"
                else:
                    error = "There is an error with the Ski/Board column. Please make sure all values are either Ski or Board."
                    return (error, status.HTTP_400_BAD_REQUEST)

                if (
                    comboDf["BOOT_IMAGE_URL"].count() > 0
                    and comboDf["BINDING_IMAGE_URL"].count() > 0
                ):
                    total = 3
                    skiBoard = comboDf["MAIN_IMAGE_URL"][0]
                    boot = comboDf["BOOT_IMAGE_URL"][0]
                    binding = comboDf["BINDING_IMAGE_URL"][0]
                    if packageType == "Ski":
                        packageImage = fn.skiPackageBuilder(skiBoard, boot, binding)
                    elif packageType == "Board":
                        packageImage = fn.boardPackageBuilder(skiBoard, boot, binding)

                elif (
                    comboDf["BOOT_IMAGE_URL"].count() > 0
                    and comboDf["BINDING_IMAGE_URL"].count() == 0
                ):
                    total = 2
                    binding = ""
                    skiBoard = comboDf["MAIN_IMAGE_URL"][0]
                    boot = comboDf["BOOT_IMAGE_URL"][0]
                    if packageType == "Ski":
                        packageImage = fn.twoItemSkiPackageBuilder(skiBoard, boot)
                    elif packageType == "Board":
                        packageImage = fn.twoItemBoardPackageBuilder(skiBoard, boot)
                elif (
                    comboDf["BOOT_IMAGE_URL"].count() == 0
                    and comboDf["BINDING_IMAGE_URL"].count() > 0
                ):
                    total = 2
                    boot = ""
                    skiBoard = comboDf["MAIN_IMAGE_URL"][0]
                    binding = comboDf["BINDING_IMAGE_URL"][0]
                    if packageType == "Ski":
                        packageImage = fn.twoItemSkiPackageBuilder(skiBoard, binding)
                    elif packageType == "Board":
                        packageImage = fn.twoItemBoardPackageBuilder(skiBoard, binding)
                elif (
                    comboDf["BOOT_IMAGE_URL"].count() == 0
                    and comboDf["BINDING_IMAGE_URL"].count() == 0
                ):
                    total = 1
                    ski = comboDf["MAIN_IMAGE_URL"][0]
                    packageImage = fn.singleSkiFileBuilder(ski)

                image_io = BytesIO()
                packageImage.convert("RGB").save(image_io, "JPEG")

                # Upload the image to the server
                image_io.seek(0)  # Reset the file pointer to the beginning

                imageNumber = 1
                folder_name = datetime.today().strftime("%Y-%m-%d")
                server_path = (
                    f"/var/www/images/media/L9/{folder_name}/{sku}_Img{imageNumber}.jpg"
                )
                with open(server_path, "wb") as f:
                    f.write(image_io.getvalue())
                BikeWagonUrl = f"https://l9golf.com/images/media/L9/{folder_name}/{sku}_Img{imageNumber}.jpg"
                df.loc[
                    df["VARIATION_PARENT_SKU"] == combo,
                    "Server Image 1",
                ] = BikeWagonUrl
                totalUploaded += 1

                # add the first image to the urlList that is used to download the csv
                urlList = BikeWagonUrl

                if skiBoard.startswith("https://l9golf.com"):
                    BikeWagonUrl = skiBoard
                else:
                    server_path = (
                        f"/var/www/images/media/L9/{folder_name}/{sku}_Img2.jpg"
                    )
                    headers = {
                        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
                    }
                    response = requests.get(skiBoard, stream=True, headers=headers)
                    image = Image.open(BytesIO(response.content)).convert("RGBA")
                    image_io = fn.process_image(image)
                    with open(server_path, "wb") as f:
                        f.write(image_io.getvalue())
                    BikeWagonUrl = f"https://l9golf.com/images/media/L9/{folder_name}/{sku}_Img2.jpg"

                df.loc[
                    df["VARIATION_PARENT_SKU"] == combo,
                    "Server Image 2",
                ] = BikeWagonUrl
                totalUploaded += 1

                urlList = urlList + "," + BikeWagonUrl

                if total == 3:
                    if boot.startswith("https://l9golf.com"):
                        BikeWagonUrl = boot
                    else:
                        server_path = (
                            f"/var/www/images/media/L9/{folder_name}/{sku}_Img3.jpg"
                        )
                        headers = {
                            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
                        }
                        response = requests.get(boot, stream=True, headers=headers)
                        image = Image.open(BytesIO(response.content)).convert("RGBA")
                        image_io = fn.process_image(image)
                        with open(server_path, "wb") as f:
                            f.write(image_io.getvalue())
                        BikeWagonUrl = f"https://l9golf.com/images/media/L9/{folder_name}/{sku}_Img3.jpg"
                    df.loc[
                        df["VARIATION_PARENT_SKU"] == combo,
                        "Server Image 3",
                    ] = BikeWagonUrl
                    totalUploaded += 1

                    urlList = urlList + "," + BikeWagonUrl

                    if binding.startswith("https://l9golf.com"):
                        BikeWagonUrl = binding
                    else:
                        server_path = (
                            f"/var/www/images/media/L9/{folder_name}/{sku}_Img4.jpg"
                        )
                        headers = {
                            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
                        }
                        response = requests.get(binding, stream=True, headers=headers)
                        image = Image.open(BytesIO(response.content)).convert("RGBA")
                        image_io = fn.process_image(image)
                        with open(server_path, "wb") as f:
                            f.write(image_io.getvalue())
                        BikeWagonUrl = f"https://l9golf.com/images/media/L9/{folder_name}/{sku}_Img4.jpg"
                    df.loc[
                        df["VARIATION_PARENT_SKU"] == combo,
                        "Server Image 4",
                    ] = BikeWagonUrl
                    totalUploaded += 1

                    urlList = urlList + "," + BikeWagonUrl
                else:
                    if boot != "":
                        if boot.startswith("https://l9golf.com"):
                            BikeWagonUrl = boot
                        else:
                            server_path = (
                                f"/var/www/images/media/L9/{folder_name}/{sku}_Img3.jpg"
                            )
                            headers = {
                                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
                            }
                            response = requests.get(boot, stream=True, headers=headers)
                            image = Image.open(BytesIO(response.content)).convert(
                                "RGBA"
                            )
                            image_io = fn.process_image(image)
                            with open(server_path, "wb") as f:
                                f.write(image_io.getvalue())
                            BikeWagonUrl = f"https://l9golf.com/images/media/L9/{folder_name}/{sku}_Img3.jpg"
                        df.loc[
                            df["VARIATION_PARENT_SKU"] == combo,
                            "Server Image 3",
                        ] = BikeWagonUrl

                        totalUploaded += 1

                        urlList = urlList + "," + BikeWagonUrl

                    elif binding != "":
                        if binding.startswith("https://l9golf.com"):
                            BikeWagonUrl = binding
                        else:
                            server_path = (
                                f"/var/www/images/media/L9/{folder_name}/{sku}_Img3.jpg"
                            )
                            headers = {
                                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36",
                            }
                            response = requests.get(
                                binding, stream=True, headers=headers
                            )
                            image = Image.open(BytesIO(response.content)).convert(
                                "RGBA"
                            )
                            image_io = fn.process_image(image)
                            with open(server_path, "wb") as f:
                                f.write(image_io.getvalue())
                            BikeWagonUrl = f"https://l9golf.com/images/media/L9/{folder_name}/{sku}_Img3.jpg"
                        df.loc[
                            df["VARIATION_PARENT_SKU"] == combo,
                            "Server Image 3",
                        ] = BikeWagonUrl
                        totalUploaded += 1

                        urlList = urlList + "," + BikeWagonUrl

                df.loc[df["VARIATION_PARENT_SKU"] == combo, "Picture URLs"] = urlList
                progress = totalUploaded / totalImages
                print(totalUploaded)
                if progress == 1:
                    progress = 0.99
                print(progress)
                update_task_field(
                    task_id=task_id,
                    field="progress",
                    value=progress,
                )
        except Exception as e:
            error = "Error creating package"
            print(e)
            return error, status.HTTP_500_INTERNAL_SERVER_ERROR
    except Exception as e:
        error = "Error connecting to server"
        print(e)
        return error, status.HTTP_500_INTERNAL_SERVER_ERROR

    df = df.rename(columns={"VARIATION_PARENT_SKU": "PARENT_SKU_COLOR"})
    df["PARENT_SKU"] = df["PARENT_SKU_COLOR"]
    df.dropna(subset=["Server Image 1"], inplace=True)

    df.set_index("INVENTORY_NUMBER", inplace=True)
    dfJson = df.to_json(orient="index")
    ResponseData = {"df": dfJson, "errorDict": BrokenUrlDict}
    # return ResponseData

    csv_bytes = df.to_csv().encode("utf-8")  # Encode CSV to bytes using UTF-8
    csv_buffer = BytesIO(csv_bytes)  # Wrap the bytes into a BytesIO object

    # Reset the buffer pointer to the start
    csv_buffer.seek(0)
    try:
        uploadedFilesPath = "/var/www/images/media/L9/uploadedFiles"
        if not os.path.exists(uploadedFilesPath):
            os.mkdir(uploadedFilesPath)
            app.logger.info("Created new folder")
        with open(f"{uploadedFilesPath}/{task_id}_1.csv", "wb") as f:
            f.write(csv_buffer.getvalue())
    except:
        print("error")

    update_task_field(task_id=task_id, field="progress", value=1)
    return


def removeBackground(image):
    from app import app
    from PIL import Image
    from io import BytesIO
    import rembg

    app.logger.info("Removing background from image")
    try:
        try:
            # remove background from image
            image = rembg.remove(image)
            app.logger.info("Background removed successfully")
        except Exception as e:
            print(f"Error: {str(e)}")
            error = "There was an error removing the background"
            json_error = {"error": error}
            app.logger.error("Error removing background")
            return json_error

        # creates a file like object to save the image to
        image_io = BytesIO()
        # saves the image with no background as a png
        image.save(image_io, format="PNG")

        # open it again with PIL so we can process it
        image2 = Image.open(image_io)
        h, w = image2.size
        if h > w:
            imageSize = (h, h)
        else:
            imageSize = (w, w)

        # Create a new square image with a white background
        square_image = Image.new("RGBA", (imageSize), (255, 255, 255, 0))

        # Calculate the position to center the original image on the new canvas
        position = (
            (square_image.width - image2.width) // 2,
            (square_image.height - image2.height) // 2,
        )

        # Paste the original image into the center of the square image
        square_image.paste(image2, position, image2)

        # convert the image to RGB
        square_image = square_image.convert("RGB")

        square_image = square_image.resize((1200, 1200))

        image_io2 = BytesIO()
        # Save the resized image to a file-like object
        square_image.save(image_io2, format="JPEG")

        # Upload the image to the server
        image_io2.seek(0)  # Reset the file pointer to the beginning
        return image_io2
    except Exception as e:
        print(f"Error: {str(e)}")
        error = "There was an error processing the image"
        json_error = {"error": error}
        app.logger.error("Error processing image")
        return json_error


def process_image(image):
    from PIL import Image
    from io import BytesIO

    # Determine the new size of the square image
    max_size = max(image.size[0], image.size[1])
    square_size = (max_size, max_size)

    # Create a new square image with a white background
    square_image = Image.new("RGBA", square_size, (255, 255, 255))

    # Determine the offset to place the original image in the center of the square image
    x_offset = int((square_size[0] - image.size[0]) / 2)
    y_offset = int((square_size[1] - image.size[1]) / 2)

    try:
        # Paste the original image into the center of the square image
        square_image.paste(image, (x_offset, y_offset), image)
    except Exception as e:
        print(e)
    # Resize the square image to be 1200 x 1200
    image_resized = square_image.resize((1200, 1200))

    # Save the resized image to a file-like object
    image_io = BytesIO()
    image_resized.convert("RGB").save(image_io, "JPEG")

    # Upload the image to the server
    image_io.seek(0)  # Reset the file pointer to the beginning

    return image_io


def twoItemSkiPackageBuilder(ski, boot):
    from PIL import Image

    ski = pilOpener(ski)
    boot = pilOpener(boot)

    # create a new blank white image
    package = Image.new("RGB", (1200, 1200), (255, 255, 255))

    # find the middle point of the ski image
    skiWidth, skiHeight = ski.size
    skiThird = skiWidth / 2.5

    # crop the ski image down the middle
    ski = ski.crop((skiThird, 0, skiWidth, skiHeight))
    skiWidth, skiHeight = ski.size

    # resize the cropped ski mainting the aspect ratio
    ski = ski.resize((int(skiWidth * 1200 / skiHeight), 1200))

    # paste that ski image on the left side of the package image
    package.paste(ski, (0, 0))

    # resize the boot image to 600x600
    boot = boot.resize((600, 600))

    # paste the boot image on the top right side of the package image
    package.paste(boot, (600, 300))
    return package


def twoItemBoardPackageBuilder(board, boardBindings):
    from PIL import Image

    board = pilOpener(board)
    boardBindings = pilOpener(boardBindings)

    # create a new blank white image
    package = Image.new("RGB", (1200, 1200), (255, 255, 255))

    # calculate 25% of the image width
    removal = board.size[0] * 0.25
    # get image width
    boardWidth, boardHeight = board.size

    # crop 25% of px from each side of the board
    board = board.crop((removal, 0, (boardWidth - removal), boardHeight))

    # paste that board image on the left side of the package image
    package.paste(board, (0, 0))

    # resize the boardBindings image to 600x600
    boardBindings = boardBindings.resize((600, 600))

    # paste the boardBindings image on the top right side of the package image
    package.paste(boardBindings, (600, 300))
    return package


def boardPackageBuilder(board, boardBindings, snowboardBoots):
    from PIL import Image

    board = pilOpener(board)
    snowboardBoots = pilOpener(snowboardBoots)
    boardBindings = pilOpener(boardBindings)

    # create a new blank white image
    package = Image.new("RGB", (1200, 1200), (255, 255, 255))

    # calculate 25% of the image width
    removal = board.size[0] * 0.25
    # get image width
    boardWidth, boardHeight = board.size

    # crop 25% of px from each side of the board
    board = board.crop((removal, 0, (boardWidth - removal), boardHeight))

    # paste that board image on the left side of the package image
    package.paste(board, (0, 0))

    # resize the snowboardBoots image to 600x600
    snowboardBoots = snowboardBoots.resize((600, 600))

    # paste the snowboardBoots image on the top right side of the package image
    package.paste(snowboardBoots, (600, 0))

    # resize the boardBindings image to 600x600
    boardBindings = boardBindings.resize((600, 600))

    # paste the boardBindings image on the bottom right side of the package image
    package.paste(boardBindings, (600, 600))
    return package


def skiPackageBuilder(ski, boot, binding):
    from PIL import Image

    ski = pilOpener(ski)
    boot = pilOpener(boot)
    binding = pilOpener(binding)

    # create a new blank white image
    package = Image.new("RGB", (1200, 1200), (255, 255, 255))

    # find the third point of the ski image
    skiWidth, skiHeight = ski.size
    skiThird = skiWidth / 2.5

    # crop the ski image down the middle
    ski = ski.crop((skiThird, 0, skiWidth, skiHeight))
    skiWidth, skiHeight = ski.size

    # resize the cropped ski mainting the aspect ratio
    ski = ski.resize((int(skiWidth * 1200 / skiHeight), 1200))

    # paste that ski image on the left side of the package image
    package.paste(ski, (0, 0))

    # resize the boot image to 600x600
    boot = boot.resize((600, 600))

    # paste the boot image on the top right side of the package image
    package.paste(boot, (600, 0))

    # resize the binding image to 600x600
    binding = binding.resize((600, 600))

    # paste the binding image on the bottom right side of the package image
    package.paste(binding, (600, 600))
    return package


def skiBuilder(unbuiltSki):
    from PIL import Image

    unbuiltSki = pilOpener(unbuiltSki)

    # Determine the new size of the square image
    max_size = max(unbuiltSki.size[0], unbuiltSki.size[1])
    square_size = (max_size, max_size)

    # Create a new square image with a white background
    square_image = Image.new("RGB", square_size, (255, 255, 255))

    # Determine the offset to place the original image in the center of the square image
    x_offset = int((square_size[0] - unbuiltSki.size[0]) / 2)
    y_offset = int((square_size[1] - unbuiltSki.size[1]) / 2)

    # Paste the original image into the center of the square image
    square_image.paste(unbuiltSki, (x_offset, y_offset), unbuiltSki)

    # create a new blank white image
    package = Image.new("RGB", (1200, 1200), (255, 255, 255))

    # find the horizontal middle point of the ski image
    skiWidth, skiHeight = square_image.size
    skiHorizontalMiddle = skiHeight / 2
    skiVerticalMiddle = skiWidth / 2

    # find 25% of the ski width
    skiRemoval = skiWidth * 0.38

    singleSki = square_image.crop((skiRemoval, 0, skiVerticalMiddle, skiHeight))
    singleSkiWidth, singleSkiHeight = singleSki.size

    # crop the ski image down the middle
    bottomSki = singleSki.crop(
        (0, skiHorizontalMiddle, singleSkiWidth, singleSkiHeight)
    )
    topSki = singleSki.crop((0, 0, singleSkiWidth, skiHorizontalMiddle))

    # resize the ski to so the height is 1200 but width is proportional
    ski = square_image.resize((int(skiWidth * 1200 / skiHeight), 1200))

    skiRemoval = ski.size[0] * 0.35
    # crop the extra white off the image
    ski = ski.crop((skiRemoval, 0, 1200 - skiRemoval, 1200))

    # resize the bottomSki to have height of 1100 but make sure the image is proportional
    bottomSkiWidth, bottomSkiHeight = bottomSki.size
    aspect_ratio = bottomSkiWidth / bottomSkiHeight
    new_width = int(1150 * aspect_ratio)
    bottomSki = bottomSki.resize((new_width, 1150))

    # resize the topSki to have height of 1100 with proportional width
    topSkiWidth, topSkiHeight = topSki.size
    aspect_ratio = topSkiWidth / topSkiHeight
    new_width = int(1150 * aspect_ratio)
    topSki = topSki.resize((new_width, 1150))
    topSkiWidth, topSkiHeight = topSki.size

    # paste that ski image on the left side of the package image
    package.paste(ski, (0, 0))

    # paste the topSki image on the bottom right side of the package image
    package.paste(topSki, (500, 100))

    # paste the bottomSki image on the top right side of the package image
    package.paste(bottomSki, (500 + topSkiWidth, 0))
    return package


def pilOpener(image):
    import requests
    from PIL import Image
    from io import BytesIO

    try:
        response = requests.get(image, stream=True)
        pilImage = Image.open(BytesIO(response.content)).convert("RGBA")
    except:
        pilImage = Image.open(image).convert("RGBA")

    return pilImage


def getToken(ca_refresh_token, ca_auth_token):
    import requests
    import time

    retryCount = 0

    url = "https://api.channeladvisor.com/oauth2/token"

    payload = f"grant_type=refresh_token&refresh_token={ca_refresh_token}"
    headers = {
        "Authorization": f"Basic {ca_auth_token}",
        "Content-Type": "application/x-www-form-urlencoded",
    }

    while retryCount < 6:
        response = requests.request("POST", url, headers=headers, data=payload)
        if response.status_code == 200:
            token = response.json()["access_token"]
            return token
        elif response.status_code == 429:
            time.sleep(15)
        else:
            error = f"Request failed with status code {response.status_code}"
        retryCount += 1

    return error


def caUpload(sku, imageUrl, imageNum, auth_token):
    import requests
    import time
    from app import logger

    retryCount = 0

    url = "https://api.channeladvisor.com/v1/Products"
    params = {"$filter": f"Sku eq '{sku}'", "$select": "ID"}
    headers = {
        "Authorization": f"Bearer {auth_token}",
        "Content-Type": "application/json",
    }
    while retryCount < 6:
        r = requests.get(
            url=url,
            headers=headers,
            params=params,
        )
        retryCount += 1
        try:
            data = r.json()
        except:
            data = {"value": []}
        if r.status_code == 200 and data["value"] != []:
            CaId = data["value"][0]["ID"]
            error = ""
            break
        elif r.status_code == 429:
            time.sleep(10)
            logger.info("waiting")
        else:
            error = f"Request failed with status code {r.status_code}"
    if error != "":
        return error

    retryCount = 0
    url = f"https://api.channeladvisor.com/v1/Images(ProductID={CaId},PlacementName='ITEMIMAGEURL{imageNum}')"
    payload = {"Url": imageUrl}
    while retryCount < 6:
        response = requests.put(url, headers=headers, json=payload)
        retryCount += 1
        if response.status_code == 204:
            return "success"
        elif response.status_code == 429:
            # try again in 10 seconds
            time.sleep(10)
            logger.info("429 - Waiting")
            error = f"Request failed with status code {response.status_code}"
        else:
            error = f"Request failed with status code {response.status_code}"
            logger.error(response.status_code, response.text)
    return (error, response.text)


def singleSkiFileBuilder(task_id, df, app, folder):
    df = df[df["VARIATION_PARENT_SKU"] != "Parent"]
    uniqueCombo = df["VARIATION_PARENT_SKU"].unique()
    folder_name = datetime.today().strftime("%Y-%m-%d")

    columns = []
    totalUploaded = 0
    totalImages = df.shape[0]

    try:
        if not os.path.exists(f"/var/www/images/media/L9/{folder_name}"):
            os.mkdir(f"/var/www/images/media/L9/{folder_name}")

        try:
            # getting the uniqueSku problem is you download images multiple times
            for combo in uniqueCombo:
                comboDf = df[df["VARIATION_PARENT_SKU"] == combo]
                sku = combo
                comboDf.reset_index(drop=True, inplace=True)
                packageType = comboDf["SKI/BOARD"][0].upper()
                if comboDf["SKI/BOARD"][0].upper() == "SKI":
                    packageType = "Ski"
                elif comboDf["SKI/BOARD"][0].upper() == "BOARD":
                    packageType = "Board"
                else:
                    error = "There is an error with the Ski/Board column. Please make sure all values are either Ski or Board."
                    return (error, status.HTTP_400_BAD_REQUEST)
                imagePath = comboDf["MAIN_IMAGE_URL"][0]
                try:
                    r = requests.get(imagePath, stream=True)
                except:
                    status_code = 500
                else:
                    status_code = r.status_code
                if status_code != 200:
                    # if the imagePath contains a . split the string and get everything before the .
                    if "." in imagePath:
                        fileName = imagePath.split(".")[0]
                        fileName = fileName.strip()
                    else:
                        fileName = imagePath
                        fileName = fileName.strip()

                    for file in folder:
                        imageName = file.filename.rsplit("/", 1)[-1]
                        # remove the file extenstion from the imageName
                        imageName = imageName.split(".")[0]

                        if imageName == fileName:
                            imagePath = file
                packageImage = skiBuilder(imagePath)

                image_io = BytesIO()
                packageImage.convert("RGB").save(image_io, "JPEG")

                # Upload the image to the server
                image_io.seek(0)  # Reset the file pointer to the beginning

                imageNumber = 1
                folder_name = datetime.today().strftime("%Y-%m-%d")
                server_path = (
                    f"/var/www/images/media/L9/{folder_name}/{sku}_Img{imageNumber}.jpg"
                )
                with open(server_path, "wb") as f:
                    f.write(image_io.getvalue())
                BikeWagonUrl = f"https://l9golf.com/images/media/L9/{folder_name}/{sku}_Img{imageNumber}.jpg"
                df.loc[
                    df["VARIATION_PARENT_SKU"] == combo,
                    "Server Image 1",
                ] = BikeWagonUrl
                progress = totalUploaded / totalImages
                if progress == 1:
                    progress = 0.99
                print(progress)
                update_task_field(
                    task_id=task_id,
                    field="progress",
                    value=progress,
                )
        except Exception as e:
            print(e)
    except Exception as e:
        print(e)
    df = df.rename(columns={"VARIATION_PARENT_SKU": "PARENT_SKU_COLOR"})
    df["PARENT_SKU"] = df["PARENT_SKU_COLOR"]
    df.dropna(subset=["Server Image 1"], inplace=True)

    df.set_index("PARENT_SKU_COLOR", inplace=True)
    # dfJson = df.to_json(orient="index")
    # ResponseData = {"df": dfJson}

    csv_bytes = df.to_csv().encode("utf-8")  # Encode CSV to bytes using UTF-8
    csv_buffer = BytesIO(csv_bytes)  # Wrap the bytes into a BytesIO object

    # Reset the buffer pointer to the start
    csv_buffer.seek(0)

    try:
        uploadedFilesPath = "/var/www/images/media/L9/uploadedFiles"
        if not os.path.exists(uploadedFilesPath):
            os.mkdir(uploadedFilesPath)
        with open(f"{uploadedFilesPath}/{task_id}_1.csv", "wb") as f:
            f.write(csv_buffer.getvalue())
    except:
        print("error")
    update_task_field(task_id=task_id, field="progress", value=1)
    return


@app.route("/folderStructure", methods=["GET"])
def getFolderStructure():
    if os.getenv("NEXT_API_TOKEN") == request.headers.get("Auth-Token"):
        BASE_PATH = "/var/www/images/CMS"
        # BASE_PATH = "/Users/willclayton/Downloads"
        folder = request.args.get("folder", BASE_PATH)
        folder = os.path.abspath(folder)

        # Ensure the folder stays within the BASE_PATH
        if not folder.startswith(BASE_PATH):
            return jsonify({"error": "Access denied"}), 403

        if not os.path.exists(folder):
            return jsonify({"error": "Folder not found"}), 404

        items = []
        for entry in os.scandir(folder):
            if entry.name.startswith("."):
                continue
            items.append(
                {
                    "name": entry.name,
                    "is_directory": entry.is_dir(),
                    "path": entry.path,
                }
            )

        items.sort(key=lambda x: x["is_directory"], reverse=True)

        return jsonify(items)
    else:
        return "Unauthorized", 401


@app.route("/uploadCmsImage", methods=["POST"])
def uploadCmsImage():
    app.logger.info("uploadCmsImage")
    if os.getenv("NEXT_API_TOKEN") == request.headers.get("Auth-Token"):
        imageFile = request.files["image"]
        folderPath = request.form["folderPath"]
        flag = request.form["flag"]
        fileName = imageFile.filename

        # flag = request.form["flag"] == "true"
        # creates a variable to pass to the html page to display the image and url
        server_dir = folderPath
        server_path = f"{server_dir}/{fileName}"
        relative_path = server_dir.removeprefix("/var/www/")
        BikeWagonUrl = f"https://l9golf.com/{relative_path}/{fileName}"

        if os.path.isfile(server_path) and flag == "false":
            flag = True
            error = "Duplicate Image. Would you like to overwrite the image?"

            data = {
                "error": error,
                "flag": flag,
                "displayImage": BikeWagonUrl,
            }
            return data
        else:
            if not os.path.exists(server_dir):
                os.makedirs(server_dir)
        try:
            # handle the file upload
            image = Image.open(imageFile).convert("RGB")

            # Create a BytesIO object to save the image temporarily
            image_io = BytesIO()
            image.save(image_io, format="JPEG")  # Save the image in the desired format
            image_io.seek(0)  # Reset the buffer position

            # Now save the file
            with open(server_path, "wb") as f:
                f.write(image_io.getvalue())

            data = {"displayImage": BikeWagonUrl, "flag": False}
            return data, 200
        except Exception as e:
            app.logger.error(f"Error during file upload: {e}")
            return (
                jsonify({"error": "An error occurred during the upload process"}),
                500,
            )
    else:
        return "Unauthorized", 401


@app.route("/deleteCmsImage", methods=["POST"])
def deleteCmsImage():
    app.logger.info("UrlUpload")
    if os.getenv("NEXT_API_TOKEN") == request.headers.get("Auth-Token"):
        imageUrl = request.json["url"]
        print(imageUrl)
        BASE_PATH = "/var/www/images/CMS"
        imagePath = imageUrl.removeprefix("https://l9golf.com/images/CMS")
        fullPath = BASE_PATH + imagePath
        print(fullPath)
        try:
            p = Path(fullPath)
            Path.unlink(p)
            return "success", 200
        except Exception as e:
            app.logger.error(f"Error during CMS file delete: {e}")
            return "error", 500
    else:
        return "Unauthorized", 401


BASE_PATH = "/var/www/images/CMS"  # Adjust to match the `/folderStructure` endpoint
ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png"}  # Allowed image types


@app.route("/searchCmsImage", methods=["GET"])
def searchCmsImage():
    if os.getenv("NEXT_API_TOKEN") != request.headers.get("Auth-Token"):
        return jsonify({"error": "Unauthorized"}), 401

    search_query = request.args.get("imageName", "").strip()
    if not search_query:
        return jsonify({"error": "Missing search query"}), 400

    base_path = Path(BASE_PATH)
    search_variations = {
        search_query,
        search_query.replace(" ", "_"),
        search_query.replace(" ", "-"),
    }
    matched_items = []

    try:
        for variation in search_variations:
            for entry in base_path.rglob(f"*{variation}*"):
                if entry.name.startswith("."):
                    continue

                matched_items.append(
                    {
                        "name": entry.name,
                        "is_directory": entry.is_dir(),
                        "path": str(entry.resolve()),
                    }
                )

        # Remove duplicates (in case multiple variations match the same file)
        matched_items = list({item["path"]: item for item in matched_items}.values())

        # Sort with directories first
        matched_items.sort(key=lambda x: x["is_directory"], reverse=True)

        return jsonify(matched_items)

    except Exception as e:
        return jsonify({"error": str(e)}), 400


if __name__ == "__main__":
    app.run()
